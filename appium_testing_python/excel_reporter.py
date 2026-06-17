import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Color Palette ───────────────────────────────────────────────────────────
COLORS = {
    'headerBg':     'FFE65100',   # Orange shade 900
    'headerFont':   'FFFFFFFF',
    'titleBg':      'FFF57C00',   # Orange shade 700
    'titleFont':    'FFFFFFFF',
    'passGreen':    'FFC6EFCE',
    'passFontG':    'FF006100',
    'failRed':      'FFFFC7CE',
    'failFontR':    'FF9C0006',
    'skipYellow':   'FFFFEB9C',
    'skipFontY':    'FF9C6500',
    'rowAlt':       'FFFFF3E0',   # Light orange alternating row
    'rowNormal':    'FFFFFFFF',
    'subHeader':    'FFE8E8E8',   # Light grey sub-header
    'subFont':      'FF424242',
    'borderColor':  'FFD6D6D6',
    'metaBg':       'FFFFE0B2',   # Orange shade 100
}

THIN_BORDER = Border(
    left=Side(style='thin', color=COLORS['borderColor']),
    right=Side(style='thin', color=COLORS['borderColor']),
    top=Side(style='thin', color=COLORS['borderColor']),
    bottom=Side(style='thin', color=COLORS['borderColor'])
)

def format_duration(ms):
    if ms < 1000:
        return f"{ms}ms"
    return f"{(ms / 1000):.2f}s"

def generate_excel_report(results, summary, output_path):
    # Ensure directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()
    
    # Setup Sheets
    dash = wb.active
    dash.title = "📊 Dashboard"
    dash.views.sheetView[0].showGridLines = False

    # ════════════════════════════════════════════════════════════════
    # SHEET 1: DASHBOARD
    # ════════════════════════════════════════════════════════════════
    # Column Widths
    dash.column_dimensions['A'].width = 3
    dash.column_dimensions['B'].width = 28
    dash.column_dimensions['C'].width = 22
    dash.column_dimensions['D'].width = 28
    dash.column_dimensions['E'].width = 22
    dash.column_dimensions['F'].width = 3

    # Title Banner
    dash.merge_cells("B2:E3")
    title_cell = dash["B2"]
    title_cell.value = "🍊  SMART FOOD WASTE MANAGEMENT"
    title_cell.font = Font(name='Calibri', size=16, bold=True, color=COLORS['titleFont'])
    title_cell.fill = PatternFill(start_color=COLORS['titleBg'], end_color=COLORS['titleBg'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Sub-title
    dash.merge_cells("B4:E4")
    sub_cell = dash["B4"]
    sub_cell.value = "Appium Mobile E2E Automation — Python Test Execution Report"
    sub_cell.font = Font(name='Calibri', size=11, italic=True, color='FF424242')
    sub_cell.fill = PatternFill(start_color='FFFFF3E0', end_color='FFFFF3E0', fill_type='solid')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Metrics Section Header
    dash["B6"] = "  📋  Execution Metrics"
    dash["B6"].font = Font(name='Calibri', size=11, bold=True, color=COLORS['headerFont'])
    dash["B6"].fill = PatternFill(start_color=COLORS['headerBg'], end_color=COLORS['headerBg'], fill_type='solid')
    dash["B6"].alignment = Alignment(horizontal='left', vertical='center')

    metrics = [
        ("Total Test Cases", summary['total']),
        ("✅  Passed", summary['passed']),
        ("❌  Failed", summary['failed']),
        ("⏭  Skipped", summary['skipped']),
        ("Pass Rate", f"{summary['passRate']}%"),
        ("Total Duration", format_duration(summary['durationMs'])),
        ("Avg per Test", format_duration(round(summary['durationMs'] / max(summary['total'], 1)))),
    ]

    for idx, (k, v) in enumerate(metrics):
        r = 7 + idx
        k_cell = dash[f"B{r}"]
        v_cell = dash[f"C{r}"]
        k_cell.value = k
        v_cell.value = v
        k_cell.font = Font(name='Calibri', size=10, bold=True)
        k_cell.fill = PatternFill(start_color=COLORS['metaBg'] if idx % 2 == 0 else COLORS['rowNormal'], fill_type='solid')
        k_cell.border = THIN_BORDER
        k_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Value Cell Custom Fills
        if "Passed" in k:
            v_cell.fill = PatternFill(start_color=COLORS['passGreen'], fill_type='solid')
            v_cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['passFontG'])
        elif "Failed" in k:
            has_fails = summary['failed'] > 0
            v_cell.fill = PatternFill(start_color=COLORS['failRed'] if has_fails else COLORS['passGreen'], fill_type='solid')
            v_cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['failFontR'] if has_fails else COLORS['passFontG'])
        elif "Pass Rate" in k:
            rate = float(summary['passRate'])
            if rate == 100.0:
                v_cell.fill = PatternFill(start_color=COLORS['passGreen'], fill_type='solid')
                v_cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['passFontG'])
            elif rate >= 80.0:
                v_cell.fill = PatternFill(start_color=COLORS['skipYellow'], fill_type='solid')
                v_cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['skipFontY'])
            else:
                v_cell.fill = PatternFill(start_color=COLORS['failRed'], fill_type='solid')
                v_cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['failFontR'])
        else:
            v_cell.fill = PatternFill(start_color=COLORS['rowNormal'] if idx % 2 == 0 else COLORS['rowAlt'], fill_type='solid')
            v_cell.font = Font(name='Calibri', size=10)
        
        v_cell.border = THIN_BORDER
        v_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Environment Info Header
    dash["D6"] = "  🖥️  Environment Details"
    dash["D6"].font = Font(name='Calibri', size=11, bold=True, color=COLORS['headerFont'])
    dash["D6"].fill = PatternFill(start_color=COLORS['headerBg'], end_color=COLORS['headerBg'], fill_type='solid')
    dash["D6"].alignment = Alignment(horizontal='left', vertical='center')

    env_info = [
        ("Application", "Smart Food Waste Mgmt Mobile"),
        ("Device / Platform", summary.get('deviceName', 'Android Emulator')),
        ("OS / API Version", summary.get('platformVersion', 'Android 12.0')),
        ("Test Driver / Library", "Appium Python Client 4.x"),
        ("Execution Tool", "Python 3.x + openpyxl"),
        ("Author / Runner", "Antigravity AI Automation"),
        ("Execution Date", summary['startTime']),
    ]

    for idx, (k, v) in enumerate(env_info):
        r = 7 + idx
        k_cell = dash[f"D{r}"]
        v_cell = dash[f"E{r}"]
        k_cell.value = k
        v_cell.value = v
        k_cell.font = Font(name='Calibri', size=10, bold=True)
        k_cell.fill = PatternFill(start_color=COLORS['metaBg'] if idx % 2 == 0 else COLORS['rowNormal'], fill_type='solid')
        k_cell.border = THIN_BORDER
        k_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        v_cell.value = v
        v_cell.font = Font(name='Calibri', size=10)
        v_cell.fill = PatternFill(start_color=COLORS['rowNormal'] if idx % 2 == 0 else COLORS['rowAlt'], fill_type='solid')
        v_cell.border = THIN_BORDER
        v_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Module Summary Table
    dash["B16"] = "  📦  Module-Wise Summary"
    dash["B16"].font = Font(name='Calibri', size=11, bold=True, color=COLORS['headerFont'])
    dash["B16"].fill = PatternFill(start_color=COLORS['headerBg'], end_color=COLORS['headerBg'], fill_type='solid')
    dash.merge_cells("B16:E16")
    dash["B16"].alignment = Alignment(horizontal='left', vertical='center')

    mod_headers = ["Module / Screen", "Total", "Passed", "Failed", "Pass Rate"]
    for ci, h in enumerate(mod_headers):
        cell = dash.cell(row=17, column=ci + 2)
        cell.value = h
        cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['headerFont'])
        cell.fill = PatternFill(start_color=COLORS['titleBg'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # Calculate modules
    modules = {}
    for r in results:
        mod = r['module']
        if mod not in modules:
            modules[mod] = {'total': 0, 'passed': 0, 'failed': 0}
        modules[mod]['total'] += 1
        if r['status'] == 'PASS':
            modules[mod]['passed'] += 1
        else:
            modules[mod]['failed'] += 1

    for idx, (mod, stats) in enumerate(modules.items()):
        r = 18 + idx
        rate = (stats['passed'] / stats['total']) * 100
        row_vals = [mod, stats['total'], stats['passed'], stats['failed'], f"{rate:.1f}%"]
        for ci, val in enumerate(row_vals):
            cell = dash.cell(row=r, column=ci + 2)
            cell.value = val
            cell.font = Font(name='Calibri', size=10, bold=(ci == 0))
            cell.fill = PatternFill(start_color=COLORS['metaBg'] if idx % 2 == 0 else COLORS['rowNormal'], fill_type='solid')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='left' if ci == 0 else 'center', vertical='center')

            if ci == 3 and stats['failed'] > 0:
                cell.fill = PatternFill(start_color=COLORS['failRed'], fill_type='solid')
                cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['failFontR'])
            if ci == 4:
                if rate == 100.0:
                    cell.fill = PatternFill(start_color=COLORS['passGreen'], fill_type='solid')
                    cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['passFontG'])
                elif rate >= 80.0:
                    cell.fill = PatternFill(start_color=COLORS['skipYellow'], fill_type='solid')
                    cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['skipFontY'])
                else:
                    cell.fill = PatternFill(start_color=COLORS['failRed'], fill_type='solid')
                    cell.font = Font(name='Calibri', size=10, bold=True, color=COLORS['failFontR'])

    # ════════════════════════════════════════════════════════════════
    # SHEET 2: DETAILED LOG
    # ════════════════════════════════════════════════════════════════
    detail = wb.create_sheet(title="📋 Detailed Log")
    detail.views.sheetView[0].showGridLines = True
    
    # Columns Widths
    detail.column_dimensions['A'].width = 8   # TC ID
    detail.column_dimensions['B'].width = 20  # Module
    detail.column_dimensions['C'].width = 25  # Sub-Module
    detail.column_dimensions['D'].width = 42  # Description
    detail.column_dimensions['E'].width = 38  # Expected
    detail.column_dimensions['F'].width = 38  # Actual
    detail.column_dimensions['G'].width = 12  # Status
    detail.column_dimensions['H'].width = 12  # Duration
    detail.column_dimensions['I'].width = 20  # Screenshot

    headers = [
        "TC ID", "Module", "Sub-Module", "Test Case Description",
        "Expected Result", "Actual Result", "Status", "Duration", "Screenshot"
    ]
    for ci, h in enumerate(headers):
        cell = detail.cell(row=1, column=ci + 1)
        cell.value = h
        cell.font = Font(name='Calibri', size=11, bold=True, color=COLORS['headerFont'])
        cell.fill = PatternFill(start_color=COLORS['headerBg'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    detail.row_dimensions[1].height = 25

    last_module = ""
    row_idx = 1
    for i, tc in enumerate(results):
        row_idx += 1

        # Module Header separator
        if tc['module'] != last_module:
            detail.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            sep_cell = detail.cell(row=row_idx, column=1)
            sep_cell.value = f"  📦  {tc['module']}"
            sep_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
            sep_cell.fill = PatternFill(start_color=COLORS['titleBg'], fill_type='solid')
            sep_cell.alignment = Alignment(horizontal='left', vertical='center')
            sep_cell.border = THIN_BORDER
            detail.row_dimensions[row_idx].height = 22
            last_module = tc['module']
            row_idx += 1

        detail.row_dimensions[row_idx].height = 36
        vals = [
            tc['tcId'], tc['module'], tc['subModule'], tc['description'],
            tc['expected'], tc['actual'], tc['status'], format_duration(tc['durationMs'])
        ]

        alt_fill = COLORS['rowAlt'] if row_idx % 2 == 0 else COLORS['rowNormal']

        for ci, v in enumerate(vals):
            cell = detail.cell(row=row_idx, column=ci + 1)
            cell.value = v
            cell.font = Font(name='Calibri', size=9, bold=(ci == 0))
            cell.fill = PatternFill(start_color=alt_fill, fill_type='solid')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center' if ci in [0, 1, 2, 6, 7] else 'left', vertical='center', wrap_text=(ci in [3, 4, 5]))

            if ci == 6:  # Status
                if v == 'PASS':
                    cell.value = "✅  PASS"
                    cell.fill = PatternFill(start_color=COLORS['passGreen'], fill_type='solid')
                    cell.font = Font(name='Calibri', size=9, bold=True, color=COLORS['passFontG'])
                else:
                    cell.value = "❌  FAIL"
                    cell.fill = PatternFill(start_color=COLORS['failRed'], fill_type='solid')
                    cell.font = Font(name='Calibri', size=9, bold=True, color=COLORS['failFontR'])
        
        # Add Screenshot cell
        sc_cell = detail.cell(row=row_idx, column=9)
        if tc.get('screenshot'):
            sc_cell.value = "View Screenshot 📸"
            sc_cell.hyperlink = f"../screenshots/{tc['screenshot']}"
            sc_cell.font = Font(name='Calibri', size=9, color='0000FF', underline='single')
        else:
            sc_cell.value = "N/A"
            sc_cell.font = Font(name='Calibri', size=9)
        
        sc_cell.fill = PatternFill(start_color=alt_fill, fill_type='solid')
        sc_cell.border = THIN_BORDER
        sc_cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(output_path)
    print(f"[SUCCESS] openpyxl report saved successfully to {output_path}")
